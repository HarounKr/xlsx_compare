import sys, time
from openpyxl import load_workbook, Workbook
from colorama import Fore, init
from openpyxl.worksheet.table import Table, TableStyleInfo

def load_vm_names_from_excel(file_path, column_name):
    # Charge les noms de VM à partir d'une colonne spécifique dans un fichier Excel.
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook.active
    column_index = None
    for i, col in enumerate(sheet[1], 1):
        if col.value == column_name:
            column_index = i
            break
    if column_index is None:
        raise ValueError(f"Column '{column_name}' not found in the file.")

    vm_names = []
    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
        if row[0].value is not None:
            vm_names.append(row[0].value)
        else:
            vm_names.append("null")
    return vm_names

def create_inventory_excel(vm_results, extract_file_path):
    """ Crée un fichier Excel avec les résultats de la comparaison. """
    wb = Workbook()
    ws = wb.active
    
    # Définition des noms de colonnes
    ws.append(['Nom VM', 'Nom Hôte', 'État S1'])
    
    # Ajout des données
    for vm, host, status in vm_results:
        ws.append([vm, host, status])

    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Création d'un style de tableau
    tab = Table(displayName="InventoryTable", ref=f"A1:C{len(vm_results)+1}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(extract_file_path)

def compare_vms(file1_vms, file1_hosts, file2_vms, vm_col1, host_col1, extract_file_path):
    """ Compare les VMs et affiche les résultats. """
    init(autoreset=True)
    found_vms = set(file2_vms)
    results = []

    for vm, host in zip(file1_vms, file1_hosts):
        new_host = host if host != 'null' else ''
        if vm in found_vms:
            print(Fore.GREEN + f"{vm_col1}: {vm}")
            results.append((vm, new_host, 'OK'))
        elif host in found_vms:
            print(Fore.GREEN + f"{host_col1}: {host} ({vm_col1}: {vm})")
            results.append((vm, new_host, 'OK'))
        else:
            print(Fore.RED + f"{vm_col1}: {vm} ({host_col1}: {host})")
            results.append((vm, new_host, 'NOT OK'))
        time.sleep(0.01)

    create_inventory_excel(results, extract_file_path)

if __name__ == "__main__":
    if len(sys.argv) != 7:
        print("Usage: python3 xlsx_compare.py <file1> <vm_col1> <host_col1> <file2> <vm_col2> <extract_file_path>")
        sys.exit(1)

    file1, vm_col1, host_col1, file2, vm_col2, extract_file_path = sys.argv[1:7]
    if not extract_file_path.endswith('.xlsx'):
        print("The output file must have a '.xlsx' extension.")
        sys.exit(1)
    # Charger les données des fichiers
    file1_vms = load_vm_names_from_excel(file1, vm_col1)
    file1_hosts = load_vm_names_from_excel(file1, host_col1)
    file2_vms = load_vm_names_from_excel(file2, vm_col2)
    # Comparer les données et afficher les résultats en couleur
    compare_vms(file1_vms, file1_hosts, file2_vms, vm_col1, host_col1, extract_file_path)